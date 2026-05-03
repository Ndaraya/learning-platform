"""
Scrape the SAT Prep curriculum (English + Math) from eddify.co and build
a canonical JSON file for seeding into the learning platform.

Also probes H5P endpoints for the SAT Question Bank (Phase A.5).

Usage (from learning-platform directory):
    py -3 tools/scrape_sat_curriculum.py           # live WP (requires App Password)
    py -3 tools/scrape_sat_curriculum.py --offline # xlsx + XML only, no WP auth

Reads WP_EDDIFY_USER and WP_EDDIFY_APP_PASSWORD from .env.local (not required in --offline mode)
Outputs:
    .tmp/sat_curriculum.json   -- teaching content (English + Math)
    .tmp/h5p_probe.json        -- H5P Question Bank investigation results
"""

import json
import os
import re
import sys
import time
import unicodedata
import requests

OFFLINE = '--offline' in sys.argv

# -- Env -----------------------------------------------------------------------

def load_env(path='.env.local'):
    env = {}
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, _, value = line.partition('=')
                    env[key.strip()] = value.strip()
    except FileNotFoundError:
        print(f'ERROR: {path} not found. Run from the learning-platform directory.')
        sys.exit(1)
    return env

env      = load_env()
WP_USER  = env.get('WP_EDDIFY_USER', '')
WP_PASS  = env.get('WP_EDDIFY_APP_PASSWORD', '')

if not OFFLINE and (not WP_USER or not WP_PASS):
    print('ERROR: WP_EDDIFY_USER and WP_EDDIFY_APP_PASSWORD must be set in .env.local')
    print('       (or run with --offline to use xlsx + XML only)')
    sys.exit(1)

BASE_URL  = 'https://eddify.co'
WP_REST   = f'{BASE_URL}/wp-json/wp/v2'
LP_REST   = f'{BASE_URL}/wp-json/lp/v1'
ADMIN_AJAX = f'{BASE_URL}/wp-admin/admin-ajax.php'
XML_PATH  = 'eddify.WordPress.2026-04-03.xml'
XLSX_PATH = 'EDDIFY-YouTube-Links.xlsx'
PLACEHOLDER_VIDEO = 'https://www.youtube.com/watch?v=dQw4w9WgXcQ'  # replaced by seeder instructions

# -- HTTP session with Basic Auth ----------------------------------------------

session = requests.Session()
session.auth = (WP_USER, WP_PASS)
session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/123.0'})

def verify_auth():
    print(f'Verifying credentials for {WP_USER}...')
    r = session.get(f'{WP_REST}/users/me', timeout=15)
    if r.status_code == 200:
        data = r.json()
        print(f'  Authenticated as: {data.get("name")} (roles: {data.get("roles", [])})')
        return True
    print(f'  ERROR: Auth failed ({r.status_code}): {r.text[:200]}')
    sys.exit(1)

def fetch(url, params=None, retries=3):
    for attempt in range(retries):
        try:
            r = session.get(url, params=params, timeout=30)
            r.raise_for_status()
            return r.text
        except requests.RequestException:
            if attempt == retries - 1:
                return ''
            time.sleep(2)
    return ''

# -- Title normalisation for fuzzy matching ------------------------------------

def normalise(s):
    s = s.lower().strip()
    s = unicodedata.normalize('NFKD', s)
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def best_match(title, lookup: dict, threshold=0.6):
    """Return the value in lookup whose normalised key most closely matches title."""
    from difflib import SequenceMatcher
    key = normalise(title)
    best_ratio = 0
    best_val = None
    for k, v in lookup.items():
        ratio = SequenceMatcher(None, key, k).ratio()
        if ratio > best_ratio:
            best_ratio, best_val = ratio, v
    return best_val if best_ratio >= threshold else None

# -- Video URL extraction ------------------------------------------------------

YT_PATTERNS = [
    r'youtube(?:-nocookie)?\.com/embed/([\w-]{11})',
    r'youtu\.be/([\w-]{11})',
    r'youtube\.com/watch\?v=([\w-]{11})',
]
VIMEO_PATTERN = r'vimeo\.com/(?:video/)?(\d+)'

def extract_video_url(text):
    for pat in YT_PATTERNS:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return f'https://www.youtube.com/watch?v={m.group(1)}'
    m = re.search(VIMEO_PATTERN, text, re.IGNORECASE)
    if m:
        return f'https://vimeo.com/{m.group(1)}'
    return None

# -- Load YouTube URLs from xlsx -----------------------------------------------

def load_xlsx_videos():
    """
    Returns {
      sheet_name: {
        'by_title':  {normalised_title: (original_title, url)},  # flat lookup
        'sections':  [(section_name, [(original_title, url), ...]), ...]  # ordered groups
      }
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('WARNING: openpyxl not installed — install with: pip install openpyxl')
        return {}

    wb = load_workbook(XLSX_PATH, read_only=True)
    result = {}
    for ws in wb.worksheets:
        by_title  = {}
        sections  = []
        sec_index = {}  # section_name -> index in sections list
        current_section = None
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            raw_sec = row[0] if len(row) > 0 else None
            topic   = row[1] if len(row) > 1 else None
            url     = row[2] if len(row) > 2 else None
            if raw_sec:
                current_section = str(raw_sec).strip()
                if current_section not in sec_index:
                    sec_index[current_section] = len(sections)
                    sections.append((current_section, []))
            skip = str(topic).lower() if topic else ''
            if topic and url and skip not in ('topic', 'section', 'youtube link', 'youtube'):
                orig = str(topic).strip()
                u    = str(url).strip()
                if current_section:
                    sections[sec_index[current_section]][1].append((orig, u))
                by_title[normalise(orig)] = (orig, u)
        result[ws.title.upper()] = {'by_title': by_title, 'sections': sections}
    return result

# -- Build title -> body HTML map from XML export -------------------------------

def build_body_map():
    """Parse XML, return two dicts: {normalised_title: body_html}, {wp_post_id: (title, body_html, h5p_ids)}"""
    try:
        from lxml import etree
    except ImportError:
        print('WARNING: lxml not installed — install with: pip install lxml')
        return {}, {}

    if not os.path.exists(XML_PATH):
        print(f'WARNING: {XML_PATH} not found — body HTML will be empty')
        return {}, {}

    CONTENT_NS = 'http://purl.org/rss/1.0/modules/content/'
    WP_NS      = 'http://wordpress.org/export/1.2/'

    title_map  = {}  # normalised_title -> body_html
    id_map     = {}  # wp_post_id -> (title, body_html, [h5p_ids])
    H5P_RE     = re.compile(r'\[h5p\s+id=["\']?(\d+)["\']?\]', re.IGNORECASE)

    context = etree.iterparse(XML_PATH, events=('end',), tag='item', recover=True)
    for _, elem in context:
        post_type_el = elem.find(f'{{{WP_NS}}}post_type')
        if post_type_el is None or post_type_el.text not in ('lp_lesson', 'page', 'post'):
            elem.clear()
            continue

        title_el   = elem.find('title')
        content_el = elem.find(f'{{{CONTENT_NS}}}encoded')
        post_id_el = elem.find(f'{{{WP_NS}}}post_id')

        title    = (title_el.text or '').strip()   if title_el is not None else ''
        body     = (content_el.text or '').strip() if content_el is not None else ''
        post_id  = (post_id_el.text or '').strip() if post_id_el is not None else ''
        h5p_ids  = H5P_RE.findall(body)

        if title:
            title_map[normalise(title)] = body
        if post_id:
            id_map[post_id] = (title, body, [int(x) for x in h5p_ids])

        elem.clear()

    print(f'  XML: indexed {len(title_map)} posts, {sum(len(v[2]) for v in id_map.values())} H5P shortcodes found')
    return title_map, id_map

# -- Discover LP courses on eddify.co -----------------------------------------

def discover_courses():
    """Return list of {id, title, slug, link} for all lp_course posts."""
    courses = []
    page = 1
    while True:
        text = fetch(f'{WP_REST}/lp_course', params={'per_page': 50, 'page': page})
        if not text:
            break
        try:
            items = json.loads(text)
            if not items or not isinstance(items, list):
                break
            for item in items:
                courses.append({
                    'id':    item.get('id'),
                    'title': item.get('title', {}).get('rendered', ''),
                    'slug':  item.get('slug', ''),
                    'link':  item.get('link', ''),
                })
            if len(items) < 50:
                break
            page += 1
        except (json.JSONDecodeError, AttributeError):
            break
    return courses

# -- Parse LP course page for module/lesson/quiz structure ---------------------

def parse_curriculum_extended(html):
    """
    Returns list of modules, each with lessons and quizzes.
    module = {title, order, lessons: [{order, item_id, title, item_type}]}
    item_type is 'lp_lesson' or 'lp_quiz'.
    """
    section_re = re.compile(
        r'<li[^>]*class="[^"]*course-section[^"]*"[^>]*data-section-id="\d+"[^>]*>(.*?)(?=<li[^>]*class="[^"]*course-section[^"]*"|$)',
        re.DOTALL | re.IGNORECASE
    )
    title_re = re.compile(
        r'<div[^>]*class="[^"]*course-section__title[^"]*"[^>]*>(.*?)</div>',
        re.DOTALL | re.IGNORECASE
    )
    item_re = re.compile(
        r'data-item-id="(\d+)"[^>]*data-item-type="(lp_lesson|lp_quiz)"[^>]*>.*?'
        r'course-item-title[^>]*>(.*?)</div>',
        re.DOTALL | re.IGNORECASE
    )

    def clean(s):
        s = re.sub(r'<[^>]+>', '', s).strip()
        for ent, rep in [('&amp;', '&'), ('&#8211;', '-'), ('&quot;', '"'),
                         ('&#8217;', "'"), ('&#8220;', '"'), ('&#8221;', '"')]:
            s = s.replace(ent, rep)
        return s

    modules = []
    for mod_i, sec in enumerate(section_re.finditer(html), 1):
        body = sec.group(1)
        title_m = title_re.search(body)
        if not title_m:
            continue
        mod_title = clean(title_m.group(1))

        items = []
        for item_i, item in enumerate(item_re.finditer(body), 1):
            items.append({
                'order':     item_i,
                'item_id':   item.group(1),
                'title':     clean(item.group(3)),
                'item_type': item.group(2),
            })

        if items:
            modules.append({'order': mod_i, 'title': mod_title, 'items': items})

    return modules

# -- Fetch video for a lesson from live WP -------------------------------------

def get_video_for_lesson(item_id, nonce):
    strategies = [
        lambda: session.get(f'{WP_REST}/lp_lesson/{item_id}', timeout=20),
        lambda: session.get(f'{WP_REST}/posts/{item_id}', timeout=20),
        lambda: session.get(f'{LP_REST}/load_content_via_ajax/', params={'id': item_id, 'nonce': nonce}, timeout=20),
        lambda: session.get(f'{LP_REST}/load_content_via_ajax/', params={'item_id': item_id, 'nonce': nonce}, timeout=20),
        lambda: session.post(ADMIN_AJAX, data={'action': 'lp_learn_item', 'item_id': item_id, 'nonce': nonce}, timeout=20),
    ]
    for fn in strategies:
        try:
            r = fn()
            text = r.text
            if not text or '"status":"error"' in text or '"code":"rest_no_route"' in text:
                continue
            video = extract_video_url(text)
            if video:
                return video
            try:
                data = r.json()
                if isinstance(data, dict):
                    rendered = data.get('content', {}).get('rendered', '') + json.dumps(data)
                    video = extract_video_url(rendered)
                    if video:
                        return video
            except Exception:
                pass
        except Exception:
            pass
    return None

# -- Fetch quiz questions from live WP -----------------------------------------

LP_Q_TYPE_MAP = {
    'single_choice': 'mcq',
    'multi_choice':  'mcq',
    'true_false':    'mcq',
    'fill_blanks':   'written',
    'sorting':       'written',
}

def fetch_quiz_questions(quiz_id, nonce):
    """
    Try multiple LP/WP endpoints to get questions for a quiz.
    Returns list of question dicts or [] if nothing found.
    """
    strategies = [
        lambda: session.get(f'{LP_REST}/quizzes/{quiz_id}/questions', timeout=20),
        lambda: session.get(f'{LP_REST}/quizzes/{quiz_id}', timeout=20),
        lambda: session.get(f'{WP_REST}/lp_quiz/{quiz_id}', timeout=20),
        lambda: session.get(f'{WP_REST}/posts/{quiz_id}', timeout=20),
        lambda: session.post(ADMIN_AJAX,
                             data={'action': 'lp_get_quiz_data', 'quiz_id': quiz_id, 'nonce': nonce},
                             timeout=20),
    ]

    raw_questions = None
    for fn in strategies:
        try:
            r = fn()
            if not r.ok or not r.text:
                continue
            data = r.json()
            # LP v1 quizzes/{id}/questions returns an array or data.items
            if isinstance(data, list) and data:
                raw_questions = data
                break
            if isinstance(data, dict):
                items = (data.get('items') or data.get('questions') or
                         data.get('data', {}).get('questions') if isinstance(data.get('data'), dict) else None)
                if items:
                    raw_questions = items
                    break
        except Exception:
            pass

    if not raw_questions:
        return []

    questions = []
    for q in raw_questions:
        if not isinstance(q, dict):
            continue

        q_type_raw = q.get('type', 'single_choice')
        q_type = LP_Q_TYPE_MAP.get(q_type_raw, 'mcq')

        # Extract prompt — varies by LP version
        prompt = (q.get('content') or q.get('title', {}) or q.get('question') or '')
        if isinstance(prompt, dict):
            prompt = prompt.get('rendered', '')
        prompt = re.sub(r'<[^>]+>', '', str(prompt)).strip()
        if not prompt:
            continue

        # Extract options for MCQ
        options = []
        correct_answer = None
        grading_rubric = None

        if q_type == 'mcq':
            raw_options = q.get('answers') or q.get('options') or []
            if isinstance(raw_options, list):
                for opt in raw_options:
                    if isinstance(opt, dict):
                        opt_text = re.sub(r'<[^>]+>', '', str(opt.get('content') or opt.get('title') or opt.get('text') or '')).strip()
                        if opt_text:
                            options.append(opt_text)
                            if opt.get('is_true') or opt.get('correct'):
                                correct_answer = opt_text
            elif isinstance(raw_options, dict):
                for key, opt in raw_options.items():
                    opt_text = re.sub(r'<[^>]+>', '', str(opt.get('content') or opt.get('title') or opt or '')).strip()
                    if opt_text:
                        options.append(opt_text)
                        if opt.get('is_true') or opt.get('correct') or q.get('correct_answer') == key:
                            correct_answer = opt_text

            if not options:
                q_type = 'written'

        if q_type == 'written':
            explanation = q.get('hint') or q.get('explanation') or q.get('feedback') or ''
            if isinstance(explanation, dict):
                explanation = explanation.get('rendered', '')
            grading_rubric = re.sub(r'<[^>]+>', '', str(explanation)).strip() or 'Award full points for a correct and well-explained answer.'

        entry = {
            'type':   q_type,
            'prompt': prompt,
            'points': 10,
        }
        if q_type == 'mcq':
            entry['options'] = options
            entry['correct_answer'] = correct_answer or (options[0] if options else '')
        else:
            entry['grading_rubric'] = grading_rubric

        questions.append(entry)

    return questions

# -- H5P Question Bank probe (Phase A.5) ---------------------------------------

def probe_h5p(id_map):
    """
    Enumerate H5P IDs from the XML body map, probe live endpoints,
    return a report dict.
    """
    print('\n-- Phase A.5: H5P Question Bank probe --')

    # Collect all H5P IDs and which lesson they appear in
    h5p_appearances = {}
    for post_id, (title, body, h5p_ids) in id_map.items():
        for hid in h5p_ids:
            if hid not in h5p_appearances:
                h5p_appearances[hid] = []
            h5p_appearances[hid].append(title)

    if not h5p_appearances:
        print('  No [h5p id=N] shortcodes found in XML.')
        return {'h5p_ids': [], 'probe_results': [], 'recommendation': 'no_h5p_found'}

    print(f'  Found {len(h5p_appearances)} unique H5P IDs across {sum(len(v) for v in h5p_appearances.values())} appearances')
    for hid, titles in sorted(h5p_appearances.items())[:10]:
        print(f'    id={hid}: {titles[0][:60]}')
    if len(h5p_appearances) > 10:
        print(f'    ... and {len(h5p_appearances) - 10} more')

    # Probe the first 3 distinct IDs
    probe_ids = sorted(h5p_appearances.keys())[:3]
    probe_results = []

    for hid in probe_ids:
        print(f'\n  Probing H5P id={hid} ({h5p_appearances[hid][0][:50]})...')
        result = {'id': hid, 'lesson': h5p_appearances[hid][0], 'endpoints': []}

        endpoints = [
            ('h5p_rest_v1',    'GET',  f'{BASE_URL}/wp-json/h5p/v1/contents/{hid}',              None),
            ('h5p_embed_ajax', 'GET',  f'{BASE_URL}/wp-admin/admin-ajax.php?action=h5p_embed&id={hid}', None),
            ('h5p_content_json','GET', f'{BASE_URL}/wp-content/uploads/h5p/content/{hid}/content/content.json', None),
            ('wp_rest_h5p_cpt','GET',  f'{WP_REST}/h5p_content/{hid}',                           None),
        ]

        for name, method, url, data in endpoints:
            try:
                r = session.request(method, url, json=data, timeout=15)
                text_preview = r.text[:300].replace('\n', ' ')
                structured = None
                has_questions = False

                if r.ok and r.text:
                    try:
                        structured = r.json()
                        # Check if this looks like H5P question data
                        json_str = json.dumps(structured)
                        has_questions = any(k in json_str for k in
                                           ['question', 'answers', 'params', 'questions',
                                            'single_choice', 'multi_choice'])
                    except Exception:
                        pass
                    # Also check raw HTML for question/answer patterns
                    if not has_questions:
                        has_questions = bool(re.search(
                            r'H5PIntegration|h5pIntegration|"questions"\s*:|"params"\s*:',
                            r.text, re.IGNORECASE
                        ))

                entry = {
                    'endpoint':       name,
                    'status':         r.status_code,
                    'has_content':    r.ok and bool(r.text),
                    'has_questions':  has_questions,
                    'preview':        text_preview[:200],
                }
                result['endpoints'].append(entry)
                status_str = f'ok' if r.ok else f'{r.status_code}'
                q_str = ' [HAS QUESTIONS]' if has_questions else ''
                print(f'    {name}: {status_str}{q_str}')

            except Exception as e:
                result['endpoints'].append({'endpoint': name, 'error': str(e)})
                print(f'    {name}: error ({e})')

        probe_results.append(result)

    # Determine recommendation
    any_has_questions = any(
        ep.get('has_questions')
        for r in probe_results
        for ep in r.get('endpoints', [])
    )
    recommendation = 'live_scraping' if any_has_questions else 'file_export'

    print(f'\n  Probe recommendation: {recommendation}')
    if recommendation == 'live_scraping':
        print('  -> At least one endpoint returns structured question data. Live scraping is viable.')
    else:
        print('  -> No endpoint returned structured question data.')
        print('    -> Next step: export wp-content/uploads/h5p/content/ via SFTP/SSH.')

    return {
        'h5p_ids':       sorted(h5p_appearances.keys()),
        'appearances':   {str(k): v for k, v in h5p_appearances.items()},
        'probe_results': probe_results,
        'recommendation': recommendation,
    }

# -- Main ----------------------------------------------------------------------

def main():
    os.makedirs('.tmp', exist_ok=True)

    if OFFLINE:
        print('Running in --offline mode: using xlsx + XML only (no live WP calls).')
        print('Quiz/question data will be empty — add via /admin/courses after seeding.\n')
    else:
        verify_auth()

    # -- Step 1: Load static sources ------------------------------------------
    print('\nLoading YouTube URLs from xlsx...')
    xlsx_videos = load_xlsx_videos()
    # Build flat URL lookup: normalised_title -> url  (values in by_title are (orig, url) tuples)
    all_xlsx_videos = {}  # normalised_title -> url
    for sheet_data in xlsx_videos.values():
        for norm_key, (orig, url) in sheet_data['by_title'].items():
            all_xlsx_videos[norm_key] = url
    for sheet, sheet_data in xlsx_videos.items():
        print(f'  {sheet}: {len(sheet_data["by_title"])} lesson URLs across {len(sheet_data["sections"])} sections')

    print('\nParsing XML export for body HTML and H5P IDs...')
    body_map, id_map = build_body_map()

    # -- Step 2: Discover WP courses ------------------------------------------
    if OFFLINE:
        sat_courses = []
    else:
        print('\nDiscovering LearnPress courses on eddify.co...')
        all_courses = discover_courses()
        if all_courses:
            for c in all_courses:
                print(f'  [{c["id"]}] {c["title"]} — {c["slug"]}')
        else:
            print('  WARNING: /wp-json/wp/v2/lp_course returned nothing. Falling back to xlsx-only structure.')

        sat_courses = [c for c in all_courses if 'sat' in c['slug'].lower() or 'sat' in c['title'].lower()]
        if not sat_courses and all_courses:
            sat_courses = all_courses

        print(f'\nUsing {len(sat_courses)} course(s) as SAT source:')
        for c in sat_courses:
            print(f'  {c["title"]} ({c["slug"]})')

    # -- Step 3: Scrape each WP course for curriculum structure ----------------
    all_modules_raw = []  # list of (subject_prefix, wp_modules_list, nonce)

    for course in sat_courses:
        print(f'\nFetching course page: {course["link"]}')
        html = fetch(course['link'])
        if not html:
            print(f'  WARNING: Could not fetch {course["link"]}')
            continue

        nonce_m = re.search(r'"nonce"\s*:\s*"([a-f0-9]+)"', html)
        nonce   = nonce_m.group(1) if nonce_m else ''
        print(f'  Nonce: {nonce or "not found"}')

        modules = parse_curriculum_extended(html)
        if not modules:
            print(f'  WARNING: No curriculum found in {course["link"]}')
            continue

        lesson_count = sum(1 for m in modules for it in m['items'] if it['item_type'] == 'lp_lesson')
        quiz_count   = sum(1 for m in modules for it in m['items'] if it['item_type'] == 'lp_quiz')
        print(f'  Found {len(modules)} modules, {lesson_count} lessons, {quiz_count} quizzes')

        title_lower = course['title'].lower()
        if 'english' in title_lower or 'reading' in title_lower or 'writing' in title_lower:
            prefix = 'SAT English'
        elif 'math' in title_lower:
            prefix = 'SAT Math'
        else:
            prefix = 'SAT'

        all_modules_raw.append((prefix, modules, nonce))

    # -- Step 4: If no WP courses found, build section-grouped structure from xlsx
    if not all_modules_raw:
        print('\nNo WP course data found — building section-grouped structure from xlsx.')
        # English modules: one module per section
        if 'ENGLISH' in xlsx_videos:
            for sec_name, lessons_in_sec in xlsx_videos['ENGLISH']['sections']:
                if not lessons_in_sec: continue
                items = [
                    {'order': i+1, 'item_id': None, 'title': orig, 'item_type': 'lp_lesson',
                     '_url': url}
                    for i, (orig, url) in enumerate(lessons_in_sec)
                ]
                all_modules_raw.append(
                    ('SAT English', [{'order': 1, 'title': sec_name, 'items': items}], '')
                )
            print(f'  English: {len(xlsx_videos["ENGLISH"]["sections"])} sections')
        # Math modules: one module per section
        if 'MATH' in xlsx_videos:
            for sec_name, lessons_in_sec in xlsx_videos['MATH']['sections']:
                if not lessons_in_sec: continue
                items = [
                    {'order': i+1, 'item_id': None, 'title': orig, 'item_type': 'lp_lesson',
                     '_url': url}
                    for i, (orig, url) in enumerate(lessons_in_sec)
                ]
                all_modules_raw.append(
                    ('SAT Math', [{'order': 1, 'title': sec_name, 'items': items}], '')
                )
            print(f'  Math: {len(xlsx_videos["MATH"]["sections"])} sections')

    # -- Step 5: Build final module/lesson/task/question structure -------------
    print('\nBuilding curriculum JSON...')
    output_modules = []
    module_order = 0

    for prefix, wp_modules, nonce in all_modules_raw:
        for wp_mod in wp_modules:
            module_order += 1
            mod_title = f'{prefix}: {wp_mod["title"]}'
            lessons = []
            lesson_order = 0
            item_list = wp_mod['items']

            for idx, item in enumerate(item_list):
                if item['item_type'] != 'lp_lesson':
                    continue

                lesson_order += 1
                lesson_title = item['title']
                lesson_id    = item.get('item_id')

                # Video URL: use _url if pre-set (xlsx offline path), then fuzzy xlsx match, then live WP
                video_url = item.get('_url')
                if not video_url:
                    video_url = best_match(lesson_title, all_xlsx_videos)
                if not video_url and lesson_id and nonce:
                    print(f'    [{lesson_order:2d}] {lesson_title[:50]:<50} fetching video...', end=' ', flush=True)
                    video_url = get_video_for_lesson(lesson_id, nonce)
                    print('ok' if video_url else 'no video')
                    time.sleep(0.3)
                else:
                    src = 'pre-set' if item.get('_url') else 'xlsx'
                    print(f'    [{lesson_order:2d}] {lesson_title[:50]:<50} {src} ok')

                # Body HTML from XML
                body_html = best_match(lesson_title, body_map) or ''

                # Attach the next item if it's a quiz
                tasks = []
                if idx + 1 < len(item_list) and item_list[idx + 1]['item_type'] == 'lp_quiz':
                    quiz_item = item_list[idx + 1]
                    quiz_id   = quiz_item.get('item_id')
                    print(f'      -> Quiz: {quiz_item["title"][:50]} (id={quiz_id})', end=' ', flush=True)
                    questions = fetch_quiz_questions(quiz_id, nonce) if quiz_id and nonce else []
                    print(f'{len(questions)} questions')
                    tasks.append({
                        'type':         'quiz',
                        'title':        quiz_item['title'],
                        'instructions': 'Complete the quiz to check your understanding.',
                        'order':        1,
                        'questions':    questions,
                    })

                lessons.append({
                    'title':          lesson_title,
                    'youtube_url':    video_url or PLACEHOLDER_VIDEO,
                    'description':    '',
                    'body_html':      body_html,
                    'order':          lesson_order,
                    'wp_item_id':     lesson_id,
                    'video_from_xlsx': bool(item.get('_url') or best_match(lesson_title, all_xlsx_videos)),
                    'tasks':          tasks,
                })

            # Unpaired quizzes (not immediately after a lesson) attach to last lesson
            for pos, it in enumerate(item_list):
                if it['item_type'] != 'lp_quiz': continue
                if pos > 0 and item_list[pos - 1]['item_type'] == 'lp_lesson': continue
                if not lessons: continue
                quiz_id   = it.get('item_id')
                questions = fetch_quiz_questions(quiz_id, nonce) if quiz_id and nonce else []
                lessons[-1]['tasks'].append({
                    'type':         'quiz',
                    'title':        it['title'],
                    'instructions': 'Complete the quiz to check your understanding.',
                    'order':        len(lessons[-1]['tasks']) + 1,
                    'questions':    questions,
                })

            output_modules.append({
                'title':   mod_title,
                'order':   module_order,
                'lessons': lessons,
            })

    # -- Step 7: Stats summary --------------------------------------------------
    total_lessons  = sum(len(m['lessons']) for m in output_modules)
    total_tasks    = sum(len(l['tasks']) for m in output_modules for l in m['lessons'])
    total_questions = sum(len(t['questions']) for m in output_modules for l in m['lessons'] for t in l['tasks'])
    lessons_missing_video = sum(
        1 for m in output_modules for l in m['lessons']
        if l['youtube_url'] == PLACEHOLDER_VIDEO
    )

    print(f'\n-- Teaching Content Summary --')
    print(f'  Modules:   {len(output_modules)}')
    print(f'  Lessons:   {total_lessons}')
    print(f'  Tasks:     {total_tasks}')
    print(f'  Questions: {total_questions}')
    if lessons_missing_video:
        print(f'  WARNING Lessons with placeholder video URL: {lessons_missing_video}')
        print('    -> Update these in the admin UI after seeding.')

    curriculum = {
        'course': {
            'title':       'SAT Prep',
            'slug':        'sat-prep',
            'description': 'A comprehensive, self-paced SAT preparation program covering English and Math. Each lesson includes a focused video explanation, followed by practice to reinforce the concept.',
            'published':   False,
        },
        'modules':  output_modules,
        'meta': {
            'total_modules':   len(output_modules),
            'total_lessons':   total_lessons,
            'total_tasks':     total_tasks,
            'total_questions': total_questions,
            'placeholder_video': PLACEHOLDER_VIDEO,
            'lessons_missing_video': lessons_missing_video,
        },
    }

    out_path = '.tmp/sat_curriculum.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(curriculum, f, indent=2, ensure_ascii=False)
    print(f'\nTeaching content saved to {out_path}')

    # -- Phase A.5: H5P probe ---------------------------------------------------
    h5p_report = probe_h5p(id_map)
    with open('.tmp/h5p_probe.json', 'w', encoding='utf-8') as f:
        json.dump(h5p_report, f, indent=2, ensure_ascii=False)
    print(f'H5P probe saved to .tmp/h5p_probe.json')

    print('\n-- Next steps --')
    print('1. Review .tmp/sat_curriculum.json — check module/lesson counts look right.')
    print('2. Review .tmp/h5p_probe.json — see the recommendation for Question Bank ingestion.')
    if h5p_report.get('recommendation') == 'live_scraping':
        print('   -> Live scraping is viable. Confirm, then extend scraper with QBank support.')
    else:
        print('   -> File export needed. Provide SFTP/SSH access to wp-content/uploads/h5p/content/.')
    print('3. When happy, run: py -3 tools/seed_sat_curriculum.py')


if __name__ == '__main__':
    main()
